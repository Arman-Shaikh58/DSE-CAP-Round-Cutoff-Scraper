"""
Phase 3: Generate a styled Excel report from the extracted cutoff data.

Reads cutoffs_raw.csv and creates dse_cutoffs_2025.xlsx with
college code, name, branch code, branch name, and cutoffs
across all 4 CAP rounds. Categories are auto-detected from the CSV header.
"""

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
CUTOFFS_CSV = BASE_DIR / "cutoffs_raw.csv"
OUTPUT_XLSX = BASE_DIR / "dse_cutoffs_2025.xlsx"

CAP_ROUNDS = [1, 2, 3, 4]


def detect_categories_from_csv(csv_path: Path) -> list[str]:
    """
    Read the CSV header and extract category names from columns
    matching the pattern 'cutoff_<category>'.

    Returns category names in title case (e.g. ['Open', 'Ews', 'Sebc']).
    """
    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        header = next(reader)

    categories = []
    for col in header:
        match = re.match(r"cutoff_(.+)", col)
        if match:
            # Convert back to display name (e.g. "open" -> "Open", "vj/dt" -> "Vj/Dt")
            raw = match.group(1)
            categories.append(raw)

    return categories


def run_phase3(categories: list[str] | None = None) -> Path:
    """
    Main entry point for Phase 3.

    Args:
        categories: List of category keys (lowercase, matching CSV columns).
                    If None, auto-detected from the CSV header.

    Returns the path to the generated Excel file.
    """
    # Load cutoff data
    records: list[dict] = []
    with open(CUTOFFS_CSV, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append(row)

    print(f"[Phase 3] Loaded {len(records)} cutoff records from {CUTOFFS_CSV}")

    # Auto-detect categories from CSV if not provided
    if categories is None:
        categories = detect_categories_from_csv(CUTOFFS_CSV)

    if not categories:
        raise ValueError("No cutoff categories found in CSV header!")

    # Build display names (for Excel headers)
    # e.g. "open" -> "Open", "vj/dt" -> "VJ/DT", "ews" -> "EWS"
    display_names: dict[str, str] = {}
    for cat in categories:
        display_names[cat] = cat.upper() if len(cat) <= 4 else cat.title()

    print(f"[Phase 3] Categories: {', '.join(display_names[c] for c in categories)}")

    # Pivot data: group by (college_code, college_name, branch_code, branch_name)
    # and spread cap rounds into columns
    pivot: dict[tuple, dict] = {}
    for rec in records:
        key = (
            rec["college_code"],
            rec["college_name"],
            rec["branch_code"],
            rec["branch_name"],
        )
        cap_round = int(rec["cap_round"])

        if key not in pivot:
            pivot[key] = {}

        for cat in categories:
            col_key = f"cap{cap_round}_{cat}"
            val = rec.get(f"cutoff_{cat}", "")
            pivot[key][col_key] = val if val else "-"

    print(f"[Phase 3] {len(pivot)} unique college-branch combinations")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DSE Cutoffs 2025"

    # -----------------------------------------------------------------------
    # Styles
    # -----------------------------------------------------------------------
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    subheader_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    data_font = Font(name="Calibri", size=10)
    number_font = Font(name="Calibri", size=10, bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Alternating row fills
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # -----------------------------------------------------------------------
    # Header rows
    # -----------------------------------------------------------------------
    # Row 1: Main header (merged cells for CAP rounds)
    base_cols = ["College Code", "College Name", "Branch Code", "Branch Name"]
    col = 1

    # Write base column headers (span 2 rows)
    for i, title in enumerate(base_cols, start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)

    # Write CAP round group headers
    col = len(base_cols) + 1
    for cap_round in CAP_ROUNDS:
        cell = ws.cell(row=1, column=col, value=f"CAP Round {cap_round}")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.merge_cells(
            start_row=1,
            start_column=col,
            end_row=1,
            end_column=col + len(categories) - 1,
        )
        col += len(categories)

    # Row 2: Sub-headers (category names for each round)
    col = len(base_cols) + 1
    for _cap_round in CAP_ROUNDS:
        for cat in categories:
            cell = ws.cell(row=2, column=col, value=display_names[cat])
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
            cell.alignment = center_align
            col += 1

    # -----------------------------------------------------------------------
    # Data rows
    # -----------------------------------------------------------------------
    row_num = 3
    sorted_keys = sorted(pivot.keys(), key=lambda k: (k[0], k[2]))

    for idx, key in enumerate(sorted_keys):
        college_code, college_name, branch_code, branch_name = key
        data = pivot[key]
        row_fill = even_fill if idx % 2 == 0 else odd_fill

        # Base columns
        for col_idx, value in enumerate(
            [college_code, college_name, branch_code, branch_name], start=1
        ):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 2 else center_align

        # Cutoff columns
        col = len(base_cols) + 1
        for cap_round in CAP_ROUNDS:
            for cat in categories:
                col_key = f"cap{cap_round}_{cat}"
                raw_val = data.get(col_key, "-")

                # Convert to number if possible
                try:
                    val = float(raw_val)
                    cell = ws.cell(row=row_num, column=col, value=val)
                    # Values are like 94.11, not 0.9411
                    cell.number_format = "0.00"
                    cell.font = number_font
                except (ValueError, TypeError):
                    cell = ws.cell(row=row_num, column=col, value=str(raw_val))
                    cell.font = data_font

                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = center_align
                col += 1

        row_num += 1

    # -----------------------------------------------------------------------
    # Auto-fit column widths
    # -----------------------------------------------------------------------
    col_widths = {
        1: 14,  # College Code
        2: 55,  # College Name
        3: 14,  # Branch Code
        4: 40,  # Branch Name
    }
    # Cutoff columns
    col = len(base_cols) + 1
    for _ in CAP_ROUNDS:
        for _ in categories:
            col_widths[col] = 10
            col += 1

    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze panes: keep headers and base columns visible
    ws.freeze_panes = "E3"

    # -----------------------------------------------------------------------
    # Title row at the very top (insert above)
    # -----------------------------------------------------------------------
    ws.insert_rows(1)
    total_cols = len(base_cols) + len(CAP_ROUNDS) * len(categories)
    title_cell = ws.cell(
        row=1,
        column=1,
        value="DSE 2025 — CAP Round Cutoffs (CS-Related Branches)",
    )
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 30

    # Update freeze panes after inserting title row
    ws.freeze_panes = "E4"

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    wb.save(OUTPUT_XLSX)
    print(f"[Phase 3] ✅ Excel saved to {OUTPUT_XLSX}")

    return OUTPUT_XLSX


if __name__ == "__main__":
    run_phase3()
